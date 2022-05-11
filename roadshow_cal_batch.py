# -*- coding: utf-8 -*-
"""
Created on Wed May  5 16:06:05 2021

@author: CaoHayashi
"""
import jieba
import jieba.analyse
from gensim.models import Word2Vec
from gensim import corpora
from gensim.similarities import SparseTermSimilarityMatrix
from gensim.similarities import WordEmbeddingSimilarityIndex
from openpyxl import load_workbook
from openpyxl import Workbook
import time

#载入网络路演问答数据
wb = load_workbook('roadshow_query_2020.xlsx')
sheet = wb.active

#载入训练好的模型
w2v_model = Word2Vec.load("luyan_segment_query.model")

#为计算出来的结果准备一个新的Excel
write_b = Workbook()
write_s = write_b.active

#批处理参数
cnt = 0
cnt_4save = 0




#计算问句向量与答句向量的SCM，qs为问句向量，rs为答句向量，返回的就是度量值。
def sim_cal(qs,rs):
    q_str_cut = list(jieba.cut(qs))
    r_str_cut = list(jieba.cut(rs))
    
    stopwords_txt_path = './hit_stopwords.txt'#这个需要自行上网查找下载
    stopwords = [line.strip() for line in open(stopwords_txt_path, 'r', encoding='utf-8').readlines()]
    
    q_list = [w for w in q_str_cut if w not in stopwords]
    r_list = [w for w in r_str_cut if w not in stopwords]
    
    
    documents = [q_list, r_list]
    dictionary = corpora.Dictionary(documents)
    
    sentence_q = dictionary.doc2bow(q_list)
    sentence_r = dictionary.doc2bow(r_list)
    
    
    
    
    similarity_index = WordEmbeddingSimilarityIndex(w2v_model.wv)#w2v_model后面一定加一个.wv
    similarity_matrix = SparseTermSimilarityMatrix(similarity_index, dictionary)
    
    similarity = similarity_matrix.inner_product(sentence_q, sentence_r, normalized=(True,True))#normalized参数要传入两个值的元组
    return format(similarity, '.3f')



#以下是此代码的main部分
for row in range(3,sheet.max_row+1):
    q_str = sheet.cell(row,4).value
    r_str = sheet.cell(row,5).value
    cosine_sim = sim_cal(q_str, r_str)
    #print(str(cnt))
    cnt += 1
    cnt_4save += 1
    if cnt >= 50000:
        print("Now it's sleep time!")
        time.sleep(3600)
        cnt = 1
        write_s.cell(row,7,cosine_sim)
        write_b.save('roadshow_1014_try.xlsx')
    else:    
        print('Now CNT is:'+str(cnt))
        if float(cosine_sim) > 0.9:
            pass
        else:
            write_s.cell(row,7,cosine_sim)
            if cnt_4save >= 1000:
                write_b.save('roadshow_1014_try.xlsx')
                cnt_4save = 1
            else:
                print('Now CNT_4SAVE is:'+str(cnt_4save))
    
    
    

write_b.save('roadshow_1014_try.xlsx')




